from openpyxl import Workbook
from openpyxl.styles import Alignment

from robots.models import Robot

from datetime import datetime, timedelta


def generate_rep():
    wb = Workbook()
    end_date = datetime.now()
    start_date = end_date - timedelta(days=700)

    header = ['Модель', 'Версия', 'Количество за неделю']

    robot_models = Robot.objects.values_list('model', flat=True).distinct()
    for model in robot_models:
        ws = wb.create_sheet(model)
        ws.append(header)
        robots_created_last_week = Robot.objects.filter(model=model, created__range=[start_date, end_date])
        version_counter = {}

        for robot in robots_created_last_week:
            version = robot.version
            if version in version_counter:
                version_counter[version] += 1
            else:
                version_counter[version] = 1

        for version, count in version_counter.items():
            ws.append([model, version, count])

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')

    del wb['Sheet']
    wb.save('report.xlsx')
